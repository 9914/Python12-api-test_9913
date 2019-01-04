#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2018/12/20 19:46
# @Author   : Python12_秋
# @Email    : 793630871@qq.com
# @File     : http_reading_excel_01
# @Software : PyCharm
# @Explain  : Excel文件数据提取
import openpyxl
from openpyxl import load_workbook
class Case:
    def __init__(self):
        self.case_id = None
        self.case_name = None
        self.case_rept = None
        self.api_name = None
        self.mathod = None
        self.url =None
        self.request_data = None
        self.expected_data = None

class DoExcel:
    def __init__(self,name,config=None):
        self.name = name
        self.config = config
        try:            #对文件名进行异常判断
            self.file_name = load_workbook(filename=name)
        except FileNotFoundError as e:
            print("{} not found ,please check file path".format(name))
            raise e
    def get_case(self,sheet_name):
        sheet = self.file_name[sheet_name]      #获取工作簿
        max_row = sheet.max_row             #获取最大列数
        cese = []                       #初始化一个列表来存读取的数据
        for i in range(2,max_row+1):
            tast_cases = Case()             #实例化一个case对象，来存放测试数据
            tast_cases.case_id = sheet.cell(row=i,column=1).value           #读取第i行，第1格数据
            tast_cases.case_name = sheet.cell(row=i, column=2).value        # 读取第i行，第2格数据
            tast_cases.case_rept = sheet.cell(row=i, column=3).value        # 读取第i行，第3格数据
            tast_cases.api_name = sheet.cell(row=i, column=4).value         # 读取第i行，第4格数据
            tast_cases.mathod = sheet.cell(row=i, column=5).value           #读取第i行，第5格数据
            tast_cases.url = sheet.cell(row=i, column=6).value              #读取第i行，第6格数据
            tast_cases.request_data = sheet.cell(row=i, column=7).value    #读取第i行，第7格数据
            tast_cases.expected_data = sheet.cell(row=i, column=8).value    #读取第i行，第8格数据
            cese.append(tast_cases)
        #根据配置文件来读取那些数据
        final_date = []       #初始化一个列表来存储最后的数据
        if self.config == 'all':        #如果配置文件里面值是'all'，读取全部文件
            final_date = cese
        else:
            for i in cese:         #遍历读取到的数据
                if i.case_id in eval(self.config):      #如果读取到的数据的id,在配置文件里面
                    final_date.append(i)         #就把数据添加到final_date列表里面
        return final_date
    def get_sheet_names(self):      #获取所有的工作簿sheet列表
        return self.file_name.sheetnames

    def write_by_case_id(self,sheet_name, case_id, actual, result):
        # 根据sheet_name定位到sheet，case_id定位到行，取到当前行actual单元格，然后赋值后保存workbook
        sh_name = self.file_name[sheet_name]  # 根据sheet名称，获取sheet对象实例
        max_row = sh_name.max_row  # 获取最大列数
        for r in range(2, max_row + 1):  # 从第二行开始循环
            case_id_r = sh_name.cell(r, 1).value  # 获取第r行，第1列,也就是case_id的值
            if case_id_r == case_id:  # 判断获取到的case_id的值是否等于传入的case_id
                sh_name.cell(r, 9).value = actual  # 写入传进来的actual，到单元格里面的actual单元格
                sh_name.cell(r,10).value = result      #写入传进来的result，到单元格里面的result单元格
                self.file_name.save(filename=self.name)  # 保存文件
                break

# if __name__ == '__main__':
#     from common import http_path
#     from common.Http_config import Reading
#     from common.http_requests_01 import HttpRequest
#     config = Reading(http_path.config_path_control).get('CASE', 'button')
#     config_url = Reading(http_path.config_path_control).get('URL', 'url_date')
#     print(config)
#     excel_date = DoExcel(http_path.case_path, config)
#     ex = excel_date.get_case('register')
#     print(type(ex),ex)
#     sheet_name = excel_date.get_sheet_names()
#     case_list = ['register','login']  # 'login', 'recharge', 'withdarw', 'bidLoan''add', 'audit', 'generateRepayments',
#     # 'geneInvestsByMemberId', 'geneInvestsByLoanId', 'getFinanceLogList'
#     for sheet in sheet_name:
#         if sheet in case_list:
#             sh = excel_date.get_case(sheet)
#             print(sheet + "测试用例个数：{}".format(len(sh)))
#             for i in sh:
#                 print(i["case_id"],config_url+i["url"],i["request_data"],i["case_rept"])
#                 resp = HttpRequest(url=config_url + i["url"],date=i["request_data"], method=i["mathod"])
#                 print(resp.get_text())
#                 if i["expected_data"] == resp.get_text():
#                     print("pass")
#                     excel_date.write_by_case_id(sheet_name=sheet, case_id=i["url"],
#                                                        actual=resp.get_text(),result="PASS")
#                 else:
#                     print("failed")
#                     excel_date.write_by_case_id(sheet_name=sheet, case_id=i["url"],
#                                                 actual=resp.get_text(),result="FAILED")
